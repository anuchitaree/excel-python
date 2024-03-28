# import pandas as pd

import pandas
import openpyxl


def read_excel_pandas(path: str):
    df = pandas.read_excel(path, sheet_name="TIE&PE", header=None)
    list_of_df = df.values.tolist()
    count = 0
    for i in range(268, 277, 2):
        print(count)
        item_name = f"{list_of_df[i][2]}"
        print(item_name)
        if item_name == "nan":
            continue
        print(len(list_of_df[i]))

        # for c in range(0,len(list_of_df[i])):

        count = count + 1

    print(count)


def read_excel_openpyxl(path: str):
    # walmart_excel = pd.read_excel(path, sheet_name='TIE&PE')
    dataframe = openpyxl.load_workbook(path)

    wb = openpyxl.load_workbook(path, data_only="True")
    # wss = wb.get_sheet_names()
    # for s in wss:
    # print(s)

    ws = wb["TIE&PE"]
    print(ws)
    for r in range(49, 100):
        for c in range(1, 43):
            d = ws.cell(row=r, column=c)
            print(d.value)
    # print(d.value)
    # print(walmart_excel.iloc[0:48, 0:3])
    pass


def file_name_separate(file_name: str):
    month_list = [
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
    ]
    len_file = file_name.split("\\")

    file_n = len_file[len(len_file) - 1]

    factory_split = file_n.split("_")
    if len(factory_split) != 2:
        return False, file_name, []

    month_split = factory_split[1].split("'")
    if len(month_split) != 2:
        return False, file_name, []

    year_split = month_split[1].split(".")
    if len(year_split) != 2:
        return False, file_name, []

    factory = factory_split[0]
    month_name = month_split[0]
    year = year_split[0]

    if month_name in month_list:
        return True, file_name, [factory, month_name, year]
